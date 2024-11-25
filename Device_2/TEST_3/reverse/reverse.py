import openpyxl

def reverse_alternate_rows(file_path, output_path):
    # Load the Excel workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Iterate through the rows
    for row in range(2, 18):
        # Reverse the row content if it's an even row (2nd, 4th, etc.)
        if row % 2 != 0:
            # Get the current row values
            row_values = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
            
            # Reverse the list of values
            row_values.reverse()
            
            # Write the reversed values back to the same row
            for col, value in enumerate(row_values, start=1):
                sheet.cell(row=row, column=col, value=value)
    """
     # Iterate through the rows
    for row in range(9, sheet.max_row + 1):
        # Reverse the row content if it's an odd row (9nd, 11th, etc.)
        if row % 2 != 0:
            # Get the current row values
            row_values = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
            
            # Reverse the list of values
            row_values.reverse()
            
            # Write the reversed values back to the same row
            for col, value in enumerate(row_values, start=1):
                sheet.cell(row=row, column=col, value=value)
     """
    
    # Save the modified workbook to a new file
    workbook.save(output_path)
    print(f"Rows reversed successfully. Output saved to {output_path}")

# Usage
input_file = 'output_grouped_coordinates_25_Nov_test3.xlsx'  # Replace with your file path
output_file = 'output_reversed_Final_25_Nov_test3_update.xlsx'  # Specify the output file name
reverse_alternate_rows(input_file, output_file)

